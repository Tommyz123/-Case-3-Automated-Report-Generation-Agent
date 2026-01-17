"""
Data Extractor Module for Case 3 Automation Agent.

This module extracts data from Excel files (SDG questionnaires and impact mechanisms)
and validates the data using Pydantic models.
"""

import logging
from pathlib import Path
from typing import List, Optional
from datetime import datetime

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from .models import (
    SDGResponse,
    ImpactMechanism,
    CompanyImpactData,
    ValidationResult,
    ValidationError
)

# Configure logging
logger = logging.getLogger(__name__)


class DataExtractor:
    """
    Extracts and parses data from Excel files.

    Responsibilities:
    - Read SDG questionnaire data (146 rows × 5 columns)
    - Extract company impact mechanism data (EmergConnect, Cloudshelf, etc.)
    - Validate data schema using Pydantic models

    Data Sources:
    - SDG问卷调查_完整中文版.xlsx / SDG questionnaire (Responses).xlsx
    - Mechanisms.xlsx / 影响评估机制_完整中文版.xlsx
    """

    def __init__(self, data_dir: str = "."):
        """
        Initialize DataExtractor.

        Args:
            data_dir: Directory containing Excel data files
        """
        self.data_dir = Path(data_dir)
        logger.info(f"DataExtractor initialized with data directory: {self.data_dir}")

    # ==================== Excel Reading ====================

    def _open_excel(self, filename: str) -> openpyxl.Workbook:
        """
        Open an Excel file with error handling.

        Args:
            filename: Name of the Excel file

        Returns:
            openpyxl.Workbook object

        Raises:
            FileNotFoundError: If file doesn't exist
            PermissionError: If file cannot be accessed
        """
        file_path = self.data_dir / filename

        if not file_path.exists():
            raise FileNotFoundError(f"Excel file not found: {file_path}")

        try:
            logger.info(f"Opening Excel file: {file_path}")
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            return workbook
        except PermissionError as e:
            raise PermissionError(f"Cannot access file {file_path}: {e}")
        except Exception as e:
            raise Exception(f"Error opening Excel file {file_path}: {e}")

    def _get_worksheet(self, workbook: openpyxl.Workbook, sheet_name: str) -> Worksheet:
        """
        Get a worksheet from workbook with error handling.

        Args:
            workbook: openpyxl Workbook object
            sheet_name: Name of the worksheet

        Returns:
            Worksheet object

        Raises:
            ValueError: If worksheet doesn't exist
        """
        if sheet_name not in workbook.sheetnames:
            raise ValueError(
                f"Worksheet '{sheet_name}' not found. "
                f"Available sheets: {workbook.sheetnames}"
            )

        return workbook[sheet_name]

    # ==================== SDG Questionnaire Extraction ====================

    def extract_sdg_questionnaire(
        self,
        filename: str = "SDG问卷调查_完整中文版.xlsx",
        sheet_name: str = "Form Responses 1"
    ) -> List[SDGResponse]:
        """
        Extract SDG questionnaire data from Excel file.

        Data structure:
        - Row 1: Column headers
        - Rows 2-147: Data (146 rows)
        - Columns: 时间戳, 公司名称, 你的名字, 联合国可持续发展目标, 如何实现该目标的描述

        Args:
            filename: Excel file name
            sheet_name: Worksheet name

        Returns:
            List of SDGResponse objects

        Raises:
            FileNotFoundError: If file doesn't exist
            ValueError: If worksheet doesn't exist or data format is invalid
        """
        logger.info(f"Extracting SDG questionnaire from {filename}/{sheet_name}")

        # Open workbook and get worksheet
        workbook = self._open_excel(filename)
        worksheet = self._get_worksheet(workbook, sheet_name)

        # Extract data
        responses = []
        row_count = 0
        error_count = 0

        # Iterate through rows (skip header row)
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            try:
                # Parse row data
                # Columns: 时间戳, 公司名称, 你的名字, 联合国可持续发展目标, 如何实现该目标的描述
                if not row or all(cell is None for cell in row):
                    # Skip empty rows
                    continue

                timestamp = row[0]
                company_name = row[1]
                contact_name = row[2]
                sdg_goals = row[3]
                implementation_description = row[4]

                # Convert timestamp to datetime if it's not already
                if isinstance(timestamp, str):
                    # Try to parse string timestamp
                    timestamp = datetime.fromisoformat(timestamp.replace(' ', 'T'))
                elif not isinstance(timestamp, datetime):
                    logger.warning(f"Row {row_idx}: Invalid timestamp type: {type(timestamp)}")
                    timestamp = datetime.now()  # Fallback

                # Handle edge cases for required fields
                # Ensure company_name is not empty
                company_name_str = str(company_name).strip() if company_name else ""
                if not company_name_str:
                    company_name_str = "未知公司"  # "Unknown Company"
                    logger.warning(f"Row {row_idx}: Empty company_name, using default")

                # Ensure contact_name is not empty
                contact_name_str = str(contact_name).strip() if contact_name else ""
                if not contact_name_str:
                    contact_name_str = "未知联系人"  # "Unknown Contact"
                    logger.warning(f"Row {row_idx}: Empty contact_name, using default")

                # Ensure implementation_description meets minimum length (10 chars)
                impl_desc_str = str(implementation_description).strip() if implementation_description else ""
                if len(impl_desc_str) < 10:
                    impl_desc_str = "未提供详细的实施描述信息"  # "No detailed implementation description provided"
                    logger.warning(f"Row {row_idx}: Short/empty implementation_description, using default")

                # Create SDGResponse object (Pydantic will validate)
                response = SDGResponse(
                    timestamp=timestamp,
                    company_name=company_name_str,
                    contact_name=contact_name_str,
                    sdg_goals=str(sdg_goals) if sdg_goals else "",
                    implementation_description=impl_desc_str
                )

                responses.append(response)
                row_count += 1

            except Exception as e:
                error_count += 1
                logger.error(f"Error parsing row {row_idx}: {e}")
                # Continue processing other rows

        workbook.close()

        logger.info(
            f"Extracted {row_count} SDG responses "
            f"({error_count} errors)"
        )

        return responses

    # ==================== Impact Mechanisms Extraction ====================

    def extract_impact_mechanisms(
        self,
        filename: str = "影响评估机制_完整中文版.xlsx",
        company_name: Optional[str] = None
    ) -> List[CompanyImpactData]:
        """
        Extract impact mechanism data from Excel file.

        File structure:
        - Multiple worksheets: EmergConnect, Cloudshelf, Sparkinity, etc.
        - Each worksheet contains:
          - Rows 1-11: Key information (SDG response, alternative scenario, stakeholders)
          - Row 14+: Mechanism data (8 columns)

        Args:
            filename: Excel file name
            company_name: Optional specific company to extract (worksheet name)
                         If None, extract all company worksheets

        Returns:
            List of CompanyImpactData objects

        Raises:
            FileNotFoundError: If file doesn't exist
        """
        logger.info(f"Extracting impact mechanisms from {filename}")

        # Open workbook
        workbook = self._open_excel(filename)

        # Determine which worksheets to process
        if company_name:
            # Extract specific company
            sheet_names = [company_name]
        else:
            # Extract all company worksheets (skip Template sheet and other non-company sheets)
            sheet_names = [
                name for name in workbook.sheetnames
                if name not in ["Template sheet", "extra FBB SME interview"]
            ]

        logger.info(f"Processing worksheets: {sheet_names}")

        # Extract data from each worksheet
        companies_data = []

        for sheet_name in sheet_names:
            try:
                company_data = self._extract_company_worksheet(workbook, sheet_name)
                if company_data:
                    companies_data.append(company_data)
            except Exception as e:
                logger.error(f"Error extracting data from worksheet '{sheet_name}': {e}")
                # Continue processing other worksheets

        workbook.close()

        logger.info(f"Extracted data for {len(companies_data)} companies")

        return companies_data

    def _extract_company_worksheet(
        self,
        workbook: openpyxl.Workbook,
        sheet_name: str
    ) -> Optional[CompanyImpactData]:
        """
        Extract data from a single company worksheet.

        Args:
            workbook: openpyxl Workbook object
            sheet_name: Name of the worksheet (company name)

        Returns:
            CompanyImpactData object or None if extraction fails
        """
        logger.info(f"Extracting data from worksheet: {sheet_name}")

        worksheet = self._get_worksheet(workbook, sheet_name)

        # Extract key information area (Rows 1-11)
        sdg_questionnaire_response = self._get_cell_value(worksheet, 1, 2)  # Row 1, Column B
        alternative_scenario = self._get_cell_value(worksheet, 3, 2)  # Row 3, Column B

        # Extract stakeholders (Rows 6-11, Column B)
        stakeholders = []
        for row_idx in range(6, 12):  # Rows 6-11
            stakeholder = self._get_cell_value(worksheet, row_idx, 2)
            if stakeholder and str(stakeholder).strip():
                stakeholders.append(str(stakeholder).strip())

        # Extract mechanism data (Row 14+)
        mechanisms = self._extract_mechanisms_data(worksheet)

        # Create CompanyImpactData object
        try:
            company_data = CompanyImpactData(
                company_name=sheet_name,
                sdg_questionnaire_response=sdg_questionnaire_response,
                alternative_scenario=alternative_scenario,
                stakeholders=stakeholders,
                mechanisms=mechanisms
            )

            logger.info(
                f"Extracted {len(stakeholders)} stakeholders and "
                f"{len(mechanisms)} mechanisms for {sheet_name}"
            )

            return company_data

        except Exception as e:
            logger.error(f"Error creating CompanyImpactData for {sheet_name}: {e}")
            return None

    def _extract_mechanisms_data(self, worksheet: Worksheet) -> List[ImpactMechanism]:
        """
        Extract mechanism data from worksheet (Row 14+).

        8 columns:
        1. Stakeholder affected
        2. Mechanism
        3. Driving Variable
        4. Type of impact
        5. Positive/Negative
        6. Method
        7. Value
        8. Unit

        Args:
            worksheet: Worksheet object

        Returns:
            List of ImpactMechanism objects
        """
        mechanisms = []

        # Start from row 14 (data rows)
        for row_idx in range(14, worksheet.max_row + 1):
            # Read 8 columns (A-H)
            stakeholder_affected = self._get_cell_value(worksheet, row_idx, 1)
            mechanism = self._get_cell_value(worksheet, row_idx, 2)
            driving_variable = self._get_cell_value(worksheet, row_idx, 3)
            type_of_impact = self._get_cell_value(worksheet, row_idx, 4)
            positive_negative = self._get_cell_value(worksheet, row_idx, 5)
            method = self._get_cell_value(worksheet, row_idx, 6)
            value = self._get_cell_value(worksheet, row_idx, 7)
            unit = self._get_cell_value(worksheet, row_idx, 8)

            # Skip rows where stakeholder_affected and mechanism are both empty
            if not stakeholder_affected and not mechanism:
                continue

            # Convert value to float if possible
            if value is not None:
                try:
                    value = float(value)
                except (ValueError, TypeError):
                    logger.warning(f"Row {row_idx}: Cannot convert value to float: {value}")
                    value = None

            # Create ImpactMechanism object
            try:
                impact_mechanism = ImpactMechanism(
                    stakeholder_affected=str(stakeholder_affected) if stakeholder_affected else "",
                    mechanism=str(mechanism) if mechanism else "",
                    driving_variable=str(driving_variable) if driving_variable else None,
                    type_of_impact=str(type_of_impact) if type_of_impact else None,
                    positive_negative=str(positive_negative) if positive_negative else None,
                    method=str(method) if method else None,
                    value=value,
                    unit=str(unit) if unit else None
                )

                mechanisms.append(impact_mechanism)

            except Exception as e:
                logger.error(f"Error creating ImpactMechanism at row {row_idx}: {e}")
                # Continue processing other rows

        return mechanisms

    def _get_cell_value(self, worksheet: Worksheet, row: int, col: int) -> Optional[str]:
        """
        Get cell value with error handling.

        Args:
            worksheet: Worksheet object
            row: Row number (1-indexed)
            col: Column number (1-indexed)

        Returns:
            Cell value as string or None
        """
        try:
            cell = worksheet.cell(row=row, column=col)
            value = cell.value

            if value is None:
                return None

            # Convert to string and strip whitespace
            return str(value).strip() if str(value).strip() else None

        except Exception as e:
            logger.warning(f"Error reading cell ({row}, {col}): {e}")
            return None

    # ==================== Schema Validation ====================

    def validate_schema(
        self,
        data: object,
        data_type: str = "unknown"
    ) -> ValidationResult:
        """
        Validate data schema using Pydantic models.

        Args:
            data: Data object to validate (SDGResponse, CompanyImpactData, etc.)
            data_type: Type of data for logging

        Returns:
            ValidationResult object
        """
        errors = []
        warnings = []

        # Check if data is a Pydantic model
        if hasattr(data, 'model_validate'):
            # Data is already validated by Pydantic
            logger.info(f"Data type '{data_type}' is already validated by Pydantic")

            # Additional business rule validations
            if isinstance(data, CompanyImpactData):
                # Validate company name is not empty
                if not data.company_name or not data.company_name.strip():
                    errors.append(ValidationError(
                        field="company_name",
                        error_type="required_field",
                        message="Company name cannot be empty"
                    ))

                # Validate mechanisms list is not empty
                if not data.mechanisms:
                    errors.append(ValidationError(
                        field="mechanisms",
                        error_type="required_field",
                        message="At least one mechanism is required"
                    ))

                # Warn if stakeholders list is empty
                if not data.stakeholders:
                    warnings.append("Stakeholders list is empty")

            elif isinstance(data, SDGResponse):
                # Validate company name is not empty
                if not data.company_name or not data.company_name.strip():
                    errors.append(ValidationError(
                        field="company_name",
                        error_type="required_field",
                        message="Company name cannot be empty"
                    ))

        is_valid = len(errors) == 0

        return ValidationResult(
            is_valid=is_valid,
            errors=errors,
            warnings=warnings
        )
